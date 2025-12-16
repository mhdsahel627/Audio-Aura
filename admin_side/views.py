# ============================================
# ALL IMPORTS (CONSOLIDATED - NO DUPLICATES)
# ============================================

# Django Core
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required,user_passes_test
from django.contrib.auth.models import User
from django.contrib.sessions.models import Session
from django.views.decorators.cache import never_cache
from django.http import HttpResponse
from django.core.paginator import Paginator

# Django Database & Query
from django.db.models import (
    Q, F, Sum, Count, Avg, 
    DecimalField, ExpressionWrapper
)
from django.db.models.functions import ExtractWeek, TruncMonth

# Django Utils
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.timezone import localtime, make_naive 


# Python Standard Library
import json
from datetime import datetime, timedelta
from decimal import Decimal
from collections import defaultdict

# Third-Party Libraries
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from django.http import HttpResponse
from django.utils.dateparse import parse_date

# Local App Imports
from .forms import AdminLoginForm
from user.models import User
from orders.models import Order, OrderItem
from products.models import Product
from category.models import Category


# ============================================
# HELPER FUNCTIONS
# ============================================

def staff_required(u):
    return u.is_staff


def is_admin(user):
    return user.is_authenticated and user.is_staff


# ============================================
# ADMIN LOGIN
# ============================================

@never_cache
def admin_login(request):
    if request.user.is_authenticated and request.user.is_staff:
        return redirect('admin_dashboard')

    if request.method == 'POST':
        form = AdminLoginForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            if user.is_staff:
                login(request, user)
                messages.success(request, 'Welcome Admin! 🎉')
                return redirect('admin_dashboard')
            else:
                messages.error(request, 'You are not authorized as admin ❌')
        else:
            messages.error(request, 'Invalid credentials ❌')
    else:
        form = AdminLoginForm()

    return render(request, 'admin/login.html', {'form': form})

# ============================================
# ADMIN DASHBOARD
# ============================================
def build_period_filter(period, today):
    """Build date filter dict based on period."""
    if period == "daily":
        return {"created_at__date": today}
    elif period == "weekly":
        return {
            "created_at__date__gte": today - timedelta(days=7),
            "created_at__date__lte": today,
        }
    elif period == "monthly":
        return {
            "created_at__date__gte": today - timedelta(days=30),
            "created_at__date__lte": today,
        }
    elif period == "yearly":
        return {
            "created_at__date__gte": today - timedelta(days=365),
            "created_at__date__lte": today,
        }
    return {}


@user_passes_test(is_admin, login_url="/admin/login")
@login_required(login_url="/admin/login")
@never_cache
def admin_dashboard(request):
    today = timezone.now().date()

    # =========================
    # GLOBAL DATE FILTER (manual range)
    # =========================
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")

    global_filter = {}
    if start_date and end_date:
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d").date()
            end_dt = datetime.strptime(end_date, "%Y-%m-%d").date()
            global_filter = {
                "created_at__date__gte": start_dt,
                "created_at__date__lte": end_dt,
            }
        except ValueError:
            pass

    # =========================
    # INDEPENDENT PERIODS (dropdowns)
    # =========================
    activity_period = request.GET.get("activity_period", "monthly")
    product_period = request.GET.get("product_period", "monthly")
    category_period = request.GET.get("category_period", "monthly")
    brand_period = request.GET.get("brand_period", "monthly")

    # Adapt to your Order model: using status field
    success_states = ["DELIVERED", "CONFIRMED", "SHIPPED", "PLACED"]

    # =========================
    # SUMMARY STATS (Global Date Filter)
    # =========================
    total_customers = User.objects.filter(is_active=True).count()

    stats_filter = {"status__in": success_states}
    stats_filter.update(global_filter)

    total_orders = Order.objects.filter(**stats_filter).count()
    total_sales_amt = (
        Order.objects.filter(**stats_filter).aggregate(total=Sum("total_amount"))["total"]
        or 0
    )
    total_pending = Order.objects.filter(
        status__in=["PLACED", "CONFIRMED", "SHIPPED"], **global_filter
    ).count()

    kpi_today_filter = {"created_at__date": today, "status__in": success_states}
    kpi_today_filter.update(global_filter)
    kpi_today = (
        Order.objects.filter(**kpi_today_filter).aggregate(total=Sum("total_amount"))[
            "total"
        ]
        or 0
    )

    # =========================
    # SALES PROGRESS (Global Date Filter)
    # =========================
    sales_progress_data = [0] * 12
    sales_progress_labels = [
        "Jan",
        "Feb",
        "Mar",
        "Apr",
        "May",
        "Jun",
        "Jul",
        "Aug",
        "Sep",
        "Oct",
        "Nov",
        "Dec",
    ]

    sales_filter = {"created_at__year": today.year, "status__in": success_states}
    sales_filter.update(global_filter)

    sales_qs = (
        Order.objects.filter(**sales_filter)
        .values(m=F("created_at__month"))
        .annotate(revenue=Sum("total_amount"))
        .order_by("m")
    )

    for row in sales_qs:
        idx = row["m"] - 1
        if 0 <= idx < 12:
            sales_progress_data[idx] = float(row["revenue"] or 0) / 1000

    # =========================
    # TOP PRODUCTS (Independent Filter + global)
    # =========================
    product_filter = {"status__in": success_states}
    product_filter.update(global_filter)
    product_filter.update(build_period_filter(product_period, today))

    product_orders = Order.objects.filter(**product_filter)
    product_items = OrderItem.objects.filter(
        order_id__in=product_orders.values_list("id", flat=True)
    )

    # OrderItem has product_name and product_id, not product FK
    top_products = list(
        product_items.values(name=F("product_name"))
        .annotate(units=Sum("quantity"), revenue=Sum("line_total"))
        .order_by("-revenue")[:10]
    )

    # =========================
    # TOP CATEGORIES (Independent Filter + global)
    # =========================
    # Build mapping from product id -> category name from Product model
    pid_to_cat = dict(Product.objects.values_list("id", "category__name"))

    cat_filter = {"status__in": success_states}
    cat_filter.update(global_filter)
    cat_filter.update(build_period_filter(category_period, today))

    cat_orders = Order.objects.filter(**cat_filter)
    cat_items = OrderItem.objects.filter(
        order_id__in=cat_orders.values_list("id", flat=True)
    )

    cat_raw = list(
        cat_items.values(pid=F("product_id"))
        .annotate(units=Sum("quantity"), revenue=Sum("line_total"))
        .order_by("-revenue")
    )

    cat_acc = defaultdict(lambda: {"units": 0, "revenue": 0})
    for r in cat_raw:
        cat_name = pid_to_cat.get(r["pid"], "Unknown")
        cat_acc[cat_name]["units"] += r["units"]
        cat_acc[cat_name]["revenue"] += float(r["revenue"] or 0)

    top_categories = sorted(
        [{"name": k, **v} for k, v in cat_acc.items()],
        key=lambda x: x["revenue"],
        reverse=True,
    )[:10]

    # =========================
    # TOP BRANDS (Independent Filter + global)
    # =========================
    pid_to_brand = dict(Product.objects.values_list("id", "brand__name"))

    brand_filter = {"status__in": success_states}
    brand_filter.update(global_filter)
    brand_filter.update(build_period_filter(brand_period, today))

    brand_orders = Order.objects.filter(**brand_filter)
    brand_items = OrderItem.objects.filter(
        order_id__in=brand_orders.values_list("id", flat=True)
    )

    brand_raw = list(
        brand_items.values(pid=F("product_id"))
        .annotate(units=Sum("quantity"), revenue=Sum("line_total"))
        .order_by("-revenue")
    )

    brand_acc = defaultdict(lambda: {"units": 0, "revenue": 0})
    for r in brand_raw:
        brand_name = pid_to_brand.get(r["pid"], "Unknown")
        brand_acc[brand_name]["units"] += r["units"]
        brand_acc[brand_name]["revenue"] += float(r["revenue"] or 0)

    top_brands = sorted(
        [{"name": k, **v} for k, v in brand_acc.items()],
        key=lambda x: x["revenue"],
        reverse=True,
    )[:10]

    # =========================
    # ACTIVITY CHART (Independent + global)
    # =========================
    activity_filter = {"status__in": success_states}
    activity_filter.update(global_filter)

    if activity_period == "daily":
        # 24 hours of today
        arr = [0] * 24
        labels = [f"{i}:00" for i in range(24)]

        activity_day_filter = {**activity_filter, "created_at__date": today}
        qs = (
            Order.objects.filter(**activity_day_filter)
            .annotate(h=F("created_at__hour"))
            .values(h=F("h"))
            .annotate(cnt=Count("id"))
            .order_by("h")
        )

        for row in qs:
            idx = row["h"]
            if 0 <= idx < 24:
                arr[idx] = row["cnt"]

        activity_categories, activity_data = labels, arr

    elif activity_period == "weekly":
        # Last 7 days
        arr = [0] * 7
        labels = []
        for i in range(6, -1, -1):
            date = today - timedelta(days=i)
            labels.append(date.strftime("%b %d"))

        activity_week_filter = {
            **activity_filter,
            "created_at__date__gte": today - timedelta(days=7),
            "created_at__date__lte": today,
        }
        qs = (
            Order.objects.filter(**activity_week_filter)
            .values(d=F("created_at__date"))
            .annotate(cnt=Count("id"))
            .order_by("d")
        )

        date_map = {row["d"]: row["cnt"] for row in qs}
        for i in range(7):
            date = today - timedelta(days=6 - i)
            arr[i] = date_map.get(date, 0)

        activity_categories, activity_data = labels, arr

    elif activity_period == "monthly":
        # Last 30 days
        arr = [0] * 30
        labels = []
        for i in range(29, -1, -1):
            date = today - timedelta(days=i)
            labels.append(date.strftime("%b %d"))

        activity_month_filter = {
            **activity_filter,
            "created_at__date__gte": today - timedelta(days=30),
            "created_at__date__lte": today,
        }
        qs = (
            Order.objects.filter(**activity_month_filter)
            .values(d=F("created_at__date"))
            .annotate(cnt=Count("id"))
            .order_by("d")
        )

        date_map = {row["d"]: row["cnt"] for row in qs}
        for i in range(30):
            date = today - timedelta(days=29 - i)
            arr[i] = date_map.get(date, 0)

        activity_categories, activity_data = labels, arr

    elif activity_period == "yearly":
        # Last 12 "months" buckets (approx by 30 days each)
        arr = [0] * 12
        labels = []
        for i in range(11, -1, -1):
            date = today - timedelta(days=i * 30)
            labels.append(date.strftime("%b %Y"))

        activity_year_filter = {
            **activity_filter,
            "created_at__date__gte": today - timedelta(days=365),
            "created_at__date__lte": today,
        }
        qs = (
            Order.objects.filter(**activity_year_filter)
            .annotate(ym=TruncMonth("created_at"))
            .values(ym=F("ym"))
            .annotate(cnt=Count("id"))
            .order_by("ym")
        )

        month_map = {}
        for row in qs:
            if row["ym"]:
                month_map[row["ym"].strftime("%b %Y")] = row["cnt"]

        for i, label in enumerate(labels):
            arr[i] = month_map.get(label, 0)

        activity_categories, activity_data = labels, arr

    else:
        activity_categories, activity_data = [], []

    context = {
        "total_customers": total_customers,
        "total_orders": total_orders,
        "total_sales_amt": total_sales_amt,
        "total_pending": total_pending,
        "kpi_target": 30000,  # static target, adjust if needed
        "kpi_revenue": int(total_sales_amt or 0),
        "kpi_today": kpi_today,
        "top_products": top_products,
        "top_categories": top_categories,
        "top_brands": top_brands,
        "sales_progress_labels": json.dumps(sales_progress_labels),
        "sales_progress_data": json.dumps(sales_progress_data),
        "activity_categories": json.dumps(activity_categories),
        "activity_data": json.dumps(activity_data),
        # expose all periods to template so your selects stay in sync
        "activity_period": activity_period,
        "product_period": product_period,
        "category_period": category_period,
        "brand_period": brand_period,
        "start_date": start_date,
        "end_date": end_date,
    }

    return render(request, "admin/dashboard.html", context)
# ============================================
# SALES REPORT
# ============================================
@user_passes_test(is_admin, login_url='admin_login')
@login_required(login_url='admin_login')
@never_cache
def sales_report(request):
    start = request.GET.get("start")
    end = request.GET.get("end")
    category_filter = request.GET.get("category", "all")
    status_filter = request.GET.get("status", "all")
    price_filter = request.GET.get("price", "all")
    
    if not start:
        start = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not end:
        end = datetime.now().strftime('%Y-%m-%d')

    items = OrderItem.objects.select_related("order", "order__user").all()

    if start:
        items = items.filter(order__created_at__date__gte=parse_date(start))
    if end:
        items = items.filter(order__created_at__date__lte=parse_date(end))
    
    if category_filter != "all":
        product_ids = Product.objects.filter(
            category__name__iexact=category_filter
        ).values_list('id', flat=True)
        items = items.filter(product_id__in=product_ids)
    
    if status_filter != "all":
        status_map = {
            "delivered": "DELIVERED",
            "confirmed": "CONFIRMED",
            "placed": "PLACED",
            "cancelled": "CANCELLED",
            "returned": "RETURNED",
            "shipped": "SHIPPED"
        }
        mapped_status = status_map.get(status_filter)
        if mapped_status:
            items = items.filter(status=mapped_status)
    
    if price_filter != "all":
        if price_filter == "700-1500":
            items = items.filter(unit_price__gte=700, unit_price__lt=1500)
        elif price_filter == "1500-3000":
            items = items.filter(unit_price__gte=1500, unit_price__lt=3000)
        elif price_filter == "3000-5000":
            items = items.filter(unit_price__gte=3000, unit_price__lt=5000)
        elif price_filter == "5000-10000":
            items = items.filter(unit_price__gte=5000, unit_price__lt=10000)
        elif price_filter == "10000-20000":
            items = items.filter(unit_price__gte=10000, unit_price__lt=20000)
        elif price_filter == "20000-60000":
            items = items.filter(unit_price__gte=20000, unit_price__lte=60000)
        elif price_filter == "60000+":
            items = items.filter(unit_price__gt=60000)

    total_revenue = items.aggregate(total=Sum('line_total'))['total'] or 0
    total_orders = items.values('order_id').distinct().count()
    total_quantity = items.aggregate(total=Sum('quantity'))['total'] or 0
    avg_order_value = total_revenue / total_orders if total_orders > 0 else 0

    category_sales = {}
    product_ids = items.values_list('product_id', flat=True).distinct()
    products_dict = {
        p.id: p.category.name if p.category else "Uncategorized"
        for p in Product.objects.filter(id__in=product_ids).select_related('category')
    }
    
    for item in items:
        cat_name = products_dict.get(item.product_id, "Uncategorized")
        if cat_name not in category_sales:
            category_sales[cat_name] = {"revenue": 0, "quantity": 0}
        category_sales[cat_name]["revenue"] += float(item.line_total)
        category_sales[cat_name]["quantity"] += item.quantity

    monthly_revenue = []
    monthly_labels = []
    
    for i in range(11, -1, -1):
        month_date = datetime.now() - timedelta(days=i*30)
        month_start = month_date.replace(day=1)
        
        if i == 0:
            month_end = datetime.now()
        else:
            next_month = month_start + timedelta(days=32)
            month_end = next_month.replace(day=1) - timedelta(days=1)
        
        month_items = items.filter(
            order__created_at__date__gte=month_start.date(),
            order__created_at__date__lte=month_end.date()
        )
        
        month_total = month_items.aggregate(total=Sum('line_total'))['total'] or 0
        monthly_revenue.append(float(month_total))
        monthly_labels.append(month_start.strftime('%b'))

    ordered_items = items.select_related('order', 'order__user').order_by('-order__created_at')[:500]
    
    recent_orders = []
    for item in ordered_items:
        image_url = item.image_url if item.image_url else None
        product_display = item.product_name
        if item.variant_color:
            product_display += f" {item.variant_color}"
        if item.offer_label:
            product_display += f" ({item.offer_label})"
        
        customer_name = "Guest"
        if item.order.user:
            if item.order.user.first_name:
                customer_name = item.order.user.first_name
            else:
                customer_name = item.order.user.username
        
        recent_orders.append({
            "order_id": item.order.order_number,
            "product": product_display.strip(),
            "customer": customer_name,
            "date": item.order.created_at.strftime('%Y-%m-%d'),
            "amount": float(item.line_total),
            "quantity": item.quantity,
            "status": item.status.lower(),
            "image": image_url
        })

    categories_list = list(
        Category.objects.filter(is_active=True)
        .values_list('name', flat=True)
        .order_by('name')
    )

    context = {
        "total_revenue": float(total_revenue),
        "total_orders": total_orders,
        "total_quantity": total_quantity,
        "avg_order_value": float(avg_order_value),
        "category_sales": category_sales,
        "monthly_revenue": json.dumps(monthly_revenue),
        "monthly_labels": json.dumps(monthly_labels),
        "recent_orders": json.dumps(recent_orders),
        "start": start,
        "end": end,
        "categories_list": categories_list,
        "selected_category": category_filter,
        "selected_status": status_filter,
        "selected_price": price_filter,
    }

    return render(request, "admin/sales.html", context)


# ============================================
# EXPORT SALES EXCEL 
# ============================================
@user_passes_test(is_admin, login_url='admin_login')
@login_required(login_url='admin_login')
@never_cache
def export_sales_excel(request):
    """Export sales report to Excel with correctly formatted date column."""
    
    start = request.GET.get("start")
    end = request.GET.get("end")
    category_filter = request.GET.get("category", "all")
    status_filter = request.GET.get("status", "all")
    price_filter = request.GET.get("price", "all")

    # ---------------------------
    # BASE QUERY WITH RELATIONS
    # ---------------------------
    items = OrderItem.objects.select_related(
        "order", "order__user", "product", "product__category"
    ).all()

    # ---------------------------
    # FILTERS
    # ---------------------------
    if start:
        items = items.filter(order__created_at__date__gte=start)

    if end:
        items = items.filter(order__created_at__date__lte=end)

    if category_filter != "all":
        items = items.filter(product__category__name__iexact=category_filter)

    if status_filter != "all":
        status_map = {
            "delivered": "DELIVERED", "confirmed": "CONFIRMED",
            "placed": "PLACED", "cancelled": "CANCELLED",
            "returned": "RETURNED", "shipped": "SHIPPED"
        }
        if status_filter in status_map:
            items = items.filter(status=status_map[status_filter])

    if price_filter != "all":
        price_ranges = {
            "700-1500": (700, 1500),
            "1500-3000": (1500, 3000),
            "3000-5000": (3000, 5000),
            "5000-10000": (5000, 10000),
            "10000-20000": (10000, 20000),
            "20000-60000": (20000, 60000)
        }
        if price_filter in price_ranges:
            low, high = price_ranges[price_filter]
            items = items.filter(unit_price__gte=low, unit_price__lt=high)
        elif price_filter == "60000+":
            items = items.filter(unit_price__gt=60000)

    # ---------------------------
    # EXCEL WORKBOOK
    # ---------------------------
    wb = Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    # ---------------------------
    # HEADERS
    # ---------------------------
    headers = ["Order ID", "Product", "Customer", "Date", "Amount", "Qty", "Status"]
    ws.append(headers)

    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # ---------------------------
    # DATA ROWS
    # ---------------------------
    for item in items:
        customer_name = item.order.user.first_name if item.order.user else "Guest"

        # Format date as text (prevent #######)
        date_str = item.order.created_at.strftime('%d-%b-%Y %I:%M %p')

        ws.append([
            item.order.order_number,
            item.product_name,
            customer_name,
            date_str,
            float(item.line_total),
            item.quantity,
            item.status
        ])

        # Force TEXT format to avoid auto-conversion
        date_cell = ws.cell(row=ws.max_row, column=4)
        date_cell.number_format = '@'

    # ---------------------------
    # COLUMN WIDTHS
    # ---------------------------
    widths = {
        'A': 18, 'B': 40, 'C': 18, 'D': 40,
        'E': 15, 'F': 12, 'G': 15
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # ---------------------------
    # STYLING DATA ROWS
    # ---------------------------
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for row in range(2, ws.max_row + 1):
        for col in range(1, 8):
            cell = ws.cell(row=row, column=col)
            cell.border = thin
            cell.alignment = Alignment(horizontal="left", vertical="center")

            # Shade alternate rows
            if row % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

    # Amount Right Align
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=5).alignment = Alignment(horizontal="right")

    ws.freeze_panes = "A2"

    # ---------------------------
    # RETURN FILE
    # ---------------------------
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    response["Content-Disposition"] = f'attachment; filename="Sales_Report_{timestamp}.xlsx"'

    wb.save(response)
    return response



# ============================================
# EXPORT SALES PDF
# ============================================
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from orders.models import OrderItem
from products.models import Product
from datetime import datetime, timedelta

@user_passes_test(is_admin, login_url='admin_login')
@login_required(login_url='admin_login')
@never_cache
def export_sales_pdf(request):
    """Export sales report as PDF with proper formatting"""
    
    # Get filters from request
    start = request.GET.get("start")
    end = request.GET.get("end")
    category_filter = request.GET.get("category", "all")
    status_filter = request.GET.get("status", "all")
    price_filter = request.GET.get("price", "all")
    
    # Default to last 30 days if no dates provided
    if not start:
        start = (datetime.now() - timedelta(days=30)).strftime('%Y-%m-%d')
    if not end:
        end = datetime.now().strftime('%Y-%m-%d')

    # ✅ FIXED: Only select_related 'order' and 'order__user' (NOT 'product')
    items = OrderItem.objects.select_related("order", "order__user").all()

    if start:
        items = items.filter(order__created_at__date__gte=parse_date(start))
    if end:
        items = items.filter(order__created_at__date__lte=parse_date(end))
    
    if category_filter != "all":
        product_ids = Product.objects.filter(
            category__name__iexact=category_filter
        ).values_list('id', flat=True)
        items = items.filter(product_id__in=product_ids)
    
    if status_filter != "all":
        status_map = {
            "delivered": "DELIVERED",
            "confirmed": "CONFIRMED",
            "placed": "PLACED",
            "cancelled": "CANCELLED",
            "returned": "RETURNED",
            "shipped": "SHIPPED"
        }
        if status_filter in status_map:
            items = items.filter(status=status_map[status_filter])
    
    if price_filter != "all":
        if price_filter == "700-1500":
            items = items.filter(unit_price__gte=700, unit_price__lt=1500)
        elif price_filter == "1500-3000":
            items = items.filter(unit_price__gte=1500, unit_price__lt=3000)
        elif price_filter == "3000-5000":
            items = items.filter(unit_price__gte=3000, unit_price__lt=5000)
        elif price_filter == "5000-10000":
            items = items.filter(unit_price__gte=5000, unit_price__lt=10000)
        elif price_filter == "10000-20000":
            items = items.filter(unit_price__gte=10000, unit_price__lt=20000)
        elif price_filter == "20000-60000":
            items = items.filter(unit_price__gte=20000, unit_price__lte=60000)
        elif price_filter == "60000+":
            items = items.filter(unit_price__gt=60000)

    # Limit to 100 orders for PDF
    items = items.order_by('-order__created_at')[:100]

    # Create PDF response
    response = HttpResponse(content_type='application/pdf')
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    response['Content-Disposition'] = f'attachment; filename="Sales_Report_{timestamp}.pdf"'

    # Create PDF with landscape orientation
    doc = SimpleDocTemplate(response, pagesize=landscape(A4))
    elements = []
    styles = getSampleStyleSheet()

    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#8b5cf6'),
        spaceAfter=30,
        alignment=TA_CENTER
    )
    title = Paragraph("📊 Sales Report", title_style)
    elements.append(title)

    # Date range subtitle
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=12,
        textColor=colors.grey,
        spaceAfter=20,
        alignment=TA_CENTER
    )
    subtitle = Paragraph(f"Period: {start} to {end}", subtitle_style)
    elements.append(subtitle)
    elements.append(Spacer(1, 0.3*inch))

    # Summary statistics
    total_revenue = sum(float(item.line_total) for item in items)
    total_orders = items.count()
    
    summary_data = [
        ['Total Orders', 'Total Revenue', 'Average Order Value'],
        [
            str(total_orders),
            f"RS{total_revenue:,.2f}",
            f"RS{total_revenue/total_orders:,.2f}" if total_orders > 0 else "₹0"
        ]
    ]
    
    summary_table = Table(summary_data, colWidths=[2.5*inch, 2.5*inch, 2.5*inch])
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f3f4f6')),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 11),
        ('TOPPADDING', (0, 1), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 10),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.4*inch))

    # Orders table header
    data = [['Order ID', 'Product', 'Customer', 'Date', 'Amount', 'Qty', 'Status']]

    # Add order rows
    for item in items:
        customer_name = "Guest"
        if item.order.user:
            customer_name = item.order.user.first_name or item.order.user.username

        # Truncate long product names
        product_name = item.product_name[:30] + "..." if len(item.product_name) > 30 else item.product_name
        
        data.append([
            item.order.order_number[:15],
            product_name,
            customer_name,
            item.order.created_at.strftime('%d-%b-%y'),
            f"RS {item.line_total:,.0f}",
            str(item.quantity),
            item.status.upper()
        ])

    # Create table
    table = Table(data, colWidths=[1.3*inch, 2.2*inch, 1.2*inch, 1*inch, 1*inch, 0.6*inch, 1*inch])
    
    # Table styling
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8b5cf6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.white),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('ALIGN', (1, 1), (1, -1), 'LEFT'),
        ('ALIGN', (2, 1), (2, -1), 'LEFT'),
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),
        ('ALIGN', (4, 1), (4, -1), 'RIGHT'),
        ('ALIGN', (5, 1), (5, -1), 'CENTER'),
        ('ALIGN', (6, 1), (6, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f9fafb')]),
    ]))
    
    elements.append(table)

    # Footer
    elements.append(Spacer(1, 0.5*inch))
    footer_style = ParagraphStyle(
        'Footer',
        parent=styles['Normal'],
        fontSize=9,
        textColor=colors.grey,
        alignment=TA_CENTER
    )
    footer_text = f"Generated on {datetime.now().strftime('%d %B %Y at %H:%M')} | Audio Aura Sales Report"
    footer = Paragraph(footer_text, footer_style)
    elements.append(footer)

    # Build PDF
    doc.build(elements)
    return response


# ============================================
# USER MANAGEMENT
# ============================================
@user_passes_test(is_admin, login_url='admin_login')
@login_required(login_url='admin_login')
@never_cache
def user_management(request):
    if not request.user.is_authenticated or not request.user.is_staff:
        return redirect('admin_login')

    users = User.objects.filter(is_staff=False, is_superuser=False)

    query = request.GET.get('q')
    if query:
        users = users.filter(
            Q(username__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query) |
            Q(email__icontains=query)
        )

    status = request.GET.get("status")
    if status == "active":
        users = users.filter(is_active=True)
    elif status == "blocked":
        users = users.filter(is_active=False)

    sort = request.GET.get("sort")
    if sort == "newest":
        users = users.order_by("-date_joined")
    elif sort == "oldest":
        users = users.order_by("date_joined")
    else:
        users = users.order_by("username")

    paginator = Paginator(users, 8)
    page_number = request.GET.get("page")
    users = paginator.get_page(page_number)

    return render(request, "admin/customers.html", {"users": users})


# ============================================
# BLOCK/UNBLOCK USER
# ============================================
@user_passes_test(is_admin, login_url='admin_login')
@login_required(login_url='admin_login')
@never_cache
def block_unblock_user(request, user_id):
    if not request.user.is_authenticated or not request.user.is_staff:
        return redirect('admin_login')

    user = get_object_or_404(User, id=user_id)
    user.is_active = not user.is_active
    user.save()

    if user.is_active:
        messages.success(request, f"{user.username} has been unblocked ✅")
    else:
        sessions = Session.objects.filter(expire_date__gte=timezone.now())
        for session in sessions:
            data = session.get_decoded()
            if data.get('_auth_user_id') == str(user.id):
                session.delete()
        messages.warning(request, f"{user.username} has been blocked 🚫")

    return redirect('customers')


# ============================================
# ADMIN LOGOUT
# ============================================

def admin_logout(request):
    logout(request)
    messages.info(request, "You have been logged out successfully.")
    return redirect('admin_login')
